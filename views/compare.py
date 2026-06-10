# =============================================================================
# views/compare.py — сторінка «Порівняння періодів»: два квартали поруч.
#
# Для фінансового директора:
#   - KPI: дохід, собівартість, витрати, фінрезультат, гроші;
#   - рух грошей за статтями (на що платимо) у двох періодах;
#   - обороти по рахунках з відхиленнями (Δ, Δ%);
#   - постачальники (631) та покупці (361) по контрагентах;
#   - експорт усього звіту в один Excel.
# =============================================================================

from __future__ import annotations

import io

import altair as alt
import pandas as pd
import streamlit as st

from osv1c import compare as cmp
from osv1c import report as rep
from osv1c.periods import tag_label


# ---------------------------------------------------------------------------
# Кешовані дані
# ---------------------------------------------------------------------------

@st.cache_data(ttl=300)
def _periods() -> list[str]:
    return rep.available_periods()


@st.cache_data(ttl=300)
def _cmp_osv(pa: str, pb: str) -> pd.DataFrame:
    return cmp.compare_osv(pa, pb)


@st.cache_data(ttl=300)
def _cashflow(pa: str, pb: str) -> pd.DataFrame:
    return cmp.cashflow_items(pa, pb)


@st.cache_data(ttl=300)
def _counterparty(pa: str, pb: str, account: str) -> pd.DataFrame:
    return cmp.counterparty_compare(pa, pb, account)


# ---------------------------------------------------------------------------
# Форматування
# ---------------------------------------------------------------------------

def _n(v: float) -> str:
    """1234567.8 → '1 234 568' (порожньо для ~нуля)."""
    if v is None or pd.isna(v) or abs(v) < 0.5:
        return ""
    return f"{v:,.0f}".replace(",", " ")


def _d(v: float) -> str:
    """Підписана дельта: '+1 234 568' / '−123'."""
    if v is None or pd.isna(v) or abs(v) < 0.5:
        return ""
    s = f"{v:+,.0f}".replace(",", " ")
    return s.replace("-", "−")


def _pct(a: float, b: float) -> str:
    if abs(a) < 0.5 and abs(b) < 0.5:
        return ""
    if abs(a) < 0.5:
        return "нове"
    if abs(b) < 0.5:
        return "зникло"
    return f"{(b - a) / abs(a) * 100:+.1f}%".replace("-", "−")


_GREEN, _RED = "color:#16a34a;font-weight:600", "color:#dc2626;font-weight:600"

# кольори «сходинок» рівнів рахунку — як на сторінці ОСВ
_LEVEL_STYLE = {
    0: ("#15324f", "#ffffff", "bold"),
    1: ("#1f4e78", "#ffffff", "bold"),
    2: ("#2e6da4", "#ffffff", "normal"),
}


def _delta_css(v) -> str:
    if isinstance(v, str):
        if v.startswith("+") or v == "нове":
            return _GREEN
        if v.startswith("−") or v == "зникло":
            return _RED
    return ""


def _style_plain(df: pd.DataFrame, delta_cols: list[str]):
    """Таблиця без ієрархії: підфарбовуємо лише колонки відхилень."""
    sty = df.style
    for c in delta_cols:
        if c in df.columns:
            sty = sty.map(_delta_css, subset=[c])
    return sty


def _style_tree(df: pd.DataFrame, levels: list[int], delta_cols: list[str]):
    """Ієрархічна таблиця: заливка за рівнем + кольорові відхилення."""
    def row_css(row):
        style = _LEVEL_STYLE.get(levels[int(row.name)])
        if style is None:
            return [_delta_css(row[c]) if c in delta_cols else ""
                    for c in df.columns]
        bg, col, fw = style
        return [f"background-color:{bg};color:{col};font-weight:{fw}"] * len(row)
    return df.style.apply(row_css, axis=1)


def _h(n_rows: int) -> int:
    return min(560, 36 * n_rows + 40)


# ---------------------------------------------------------------------------
# Вибір періодів
# ---------------------------------------------------------------------------

st.title("📈 Порівняння періодів")

periods = _periods()
if len(periods) < 2:
    st.error("Для порівняння потрібно щонайменше два періоди в БД. "
             "Завантажте: `python main.py --db --period 2025-Q1 2026-Q1`")
    st.stop()

with st.sidebar:
    st.header("Періоди")
    pa = st.selectbox("База (A)", periods, index=0, format_func=tag_label)
    pb = st.selectbox("Порівняння (B)", periods, index=len(periods) - 1,
                      format_func=tag_label)
    if pa == pb:
        st.warning("Оберіть два різні періоди.")
        st.stop()
    no_internal = st.checkbox(
        "Без внутрішніх переказів", value=True,
        help="Перекази між власними рахунками однаково збільшують і "
             "надходження, і платежі — для аналізу їх краще прибрати.")

osv = _cmp_osv(pa, pb)
cash = _cashflow(pa, pb)
cash_view = cash[~cash["internal"]] if no_internal else cash

# ---------------------------------------------------------------------------
# KPI
# ---------------------------------------------------------------------------

r70 = cmp.kpi_row(osv, "70")    # доходи від реалізації
r90 = cmp.kpi_row(osv, "90")    # собівартість
r92 = cmp.kpi_row(osv, "92")    # адміністративні витрати
r93 = cmp.kpi_row(osv, "93")    # витрати на збут
r36 = cmp.kpi_row(osv, "36")    # покупці
r63 = cmp.kpi_row(osv, "63")    # постачальники
fin_a, fin_b = cmp.profit_change(osv)
money = cmp.cash_kpis(osv)
in_a, in_b = cash_view["надходження_a"].sum(), cash_view["надходження_b"].sum()
out_a, out_b = cash_view["платежі_a"].sum(), cash_view["платежі_b"].sum()


def _metric(col, label, a, b, inverse=False, help=None):
    d = b - a
    if abs(d) < 0.5:
        delta = None
    else:
        # ASCII-мінус обов'язковий: за ним Streamlit визначає напрям стрілки
        pct = _pct(a, b).replace("−", "-")
        delta = f"{d:+,.0f}".replace(",", " ") + (f" ({pct})" if pct else "")
    col.metric(label, _n(b) or "0", delta=delta,
               delta_color="inverse" if inverse else "normal", help=help)


st.caption(f"Усі показники: **{pb}** проти **{pa}** (Δ та Δ% — зміна до бази), грн.")

c1, c2, c3, c4 = st.columns(4)
_metric(c1, "Дохід (оборот Кт 70)", r70.get("об_кт_a", 0), r70.get("об_кт_b", 0))
_metric(c2, "Собівартість (оборот Дт 90)",
        r90.get("об_дт_a", 0), r90.get("об_дт_b", 0), inverse=True)
_metric(c3, "Адмін. + збут (Дт 92+93)",
        r92.get("об_дт_a", 0) + r93.get("об_дт_a", 0),
        r92.get("об_дт_b", 0) + r93.get("об_дт_b", 0), inverse=True)
_metric(c4, "Фінрезультат періоду (рах. 44)", fin_a, fin_b,
        help="Зміна нерозподіленого прибутку (сальдо 44) за період")

c5, c6, c7, c8 = st.columns(4)
_metric(c5, "Надходження грошей", in_a, in_b,
        help="Обороти Дт за рахунками 30/31/33 за статтями руху коштів")
_metric(c6, "Платежі", out_a, out_b, inverse=True,
        help="Обороти Кт за рахунками 30/31/33 за статтями руху коштів")
_metric(c7, "Гроші на кінець періоду", money["кон_a"], money["кон_b"])
_metric(c8, "Борг покупців на кінець (36)",
        r36.get("кон_нетто_a", 0), r36.get("кон_нетто_b", 0), inverse=True)

st.divider()

# ---------------------------------------------------------------------------
# Вкладки
# ---------------------------------------------------------------------------

tab_cash, tab_osv, tab_sup, tab_cust = st.tabs([
    "💰 Рух грошей (статті)", "📒 Обороти по рахунках",
    "🏭 Постачальники (631)", "🛒 Покупці (361)",
])

# --- 1. Рух грошей за статтями -------------------------------------------
with tab_cash:
    st.caption("Статті руху коштів з банківських виписок (субконто рахунків "
               "30/31/33): **на що саме** платимо і **звідки** надходять гроші.")

    cf = cash_view.copy()
    cf["Δ платежі"] = cf["платежі_b"] - cf["платежі_a"]
    cf = cf.sort_values(["платежі_b", "надходження_b"],
                        ascending=False).reset_index(drop=True)

    disp = pd.DataFrame({
        "Стаття": cf["стаття"],
        f"Платежі {pa}": cf["платежі_a"].map(_n),
        f"Платежі {pb}": cf["платежі_b"].map(_n),
        "Δ платежі": cf["Δ платежі"].map(_d),
        "Δ% платежі": [_pct(a, b) for a, b in zip(cf["платежі_a"], cf["платежі_b"])],
        f"Надходження {pa}": cf["надходження_a"].map(_n),
        f"Надходження {pb}": cf["надходження_b"].map(_n),
        "Δ надходж.": (cf["надходження_b"] - cf["надходження_a"]).map(_d),
    })
    st.dataframe(_style_plain(disp, ["Δ надходж.", "Δ платежі", "Δ% платежі"]),
                 width="stretch", hide_index=True, height=_h(len(disp)))

    top = cf[cf["платежі_a"] + cf["платежі_b"] > 0].head(10)
    if not top.empty:
        long = pd.concat([
            pd.DataFrame({"Стаття": top["стаття"], "Період": pa,
                          "Платежі": top["платежі_a"]}),
            pd.DataFrame({"Стаття": top["стаття"], "Період": pb,
                          "Платежі": top["платежі_b"]}),
        ])
        chart = (alt.Chart(long).mark_bar()
                 .encode(y=alt.Y("Стаття:N", sort="-x", title=None),
                         x=alt.X("Платежі:Q", title="Платежі, грн"),
                         yOffset="Період:N",
                         color=alt.Color("Період:N",
                                         scale=alt.Scale(range=["#94a3b8", "#1f4e78"])),
                         tooltip=["Стаття", "Період",
                                  alt.Tooltip("Платежі:Q", format=",.0f")])
                 .properties(height=380, title="Найбільші статті платежів"))
        st.altair_chart(chart, width="stretch")

# --- 2. Обороти по рахунках ------------------------------------------------
with tab_osv:
    f1, f2, f3 = st.columns([2, 2, 3])
    metric = f1.radio("Показник", ["Оборот Дт", "Оборот Кт", "Сальдо кін. (нетто)"],
                      horizontal=True)
    depth_name = f2.radio("Деталізація",
                          ["Рахунки", "Субрахунки", "Повна"], horizontal=True)
    search = f3.text_input("Пошук (код або назва)", "").strip().lower()
    with_offbal = st.checkbox("Показати забалансові рахунки", value=False)

    col_a, col_b = {
        "Оборот Дт": ("об_дт_a", "об_дт_b"),
        "Оборот Кт": ("об_кт_a", "об_кт_b"),
        "Сальдо кін. (нетто)": ("кон_нетто_a", "кон_нетто_b"),
    }[metric]
    max_level = {"Рахунки": 1, "Субрахунки": 2, "Повна": 99}[depth_name]

    sel = osv[osv["level"] <= max_level].copy()
    if not with_offbal:
        sel = sel[~sel["off_balance"]]
    sel = sel[(sel[col_a].abs() > 0.5) | (sel[col_b].abs() > 0.5)]
    if search:
        sel = sel[sel["code"].str.lower().str.contains(search)
                  | sel["name"].str.lower().str.contains(search)]
    sel = sel.reset_index(drop=True)

    indent = "        "
    disp = pd.DataFrame({
        "Рахунок": sel["code"],
        "Найменування": [indent * lv + str(nm)
                         for lv, nm in zip(sel["level"], sel["name"])],
        f"{metric} {pa}": sel[col_a].map(_n),
        f"{metric} {pb}": sel[col_b].map(_n),
        "Δ": (sel[col_b] - sel[col_a]).map(_d),
        "Δ%": [_pct(a, b) for a, b in zip(sel[col_a], sel[col_b])],
    })
    st.dataframe(_style_tree(disp, sel["level"].tolist(), ["Δ", "Δ%"]),
                 width="stretch", hide_index=True, height=_h(len(disp)))
    st.caption("Рядки-групи (класи/рахунки) включають усі свої субрахунки. "
               "Δ% — зміна відносно бази (A).")


# --- 3/4. Контрагенти -------------------------------------------------------

def _counterparty_tab(account: str, col_turn: str, col_pay: str,
                      turn_label: str, pay_label: str, debt_label: str,
                      debt_sign: float):
    """
    Таблиця по контрагентах рахунку розрахунків.
      col_turn — колонка нарахованого обороту (для 631 це Кт, для 361 — Дт);
      col_pay  — колонка оплат (для 631 — Дт, для 361 — Кт);
      debt_sign — знак нетто-сальдо, щоб борг показувався додатним.
    """
    df = _counterparty(pa, pb, account)
    if df.empty:
        st.info(f"По рахунку {account} немає деталізації у вибраних періодах.")
        return

    g1, g2, g3 = st.columns(3)
    _metric(g1, f"{turn_label} (всього)",
            df[col_turn + "_a"].sum(), df[col_turn + "_b"].sum())
    _metric(g2, f"{pay_label} (всього)",
            df[col_pay + "_a"].sum(), df[col_pay + "_b"].sum())
    _metric(g3, f"{debt_label} на кінець",
            df["нетто_a"].sum() * debt_sign, df["нетто_b"].sum() * debt_sign,
            inverse=True)

    only_active = st.checkbox("Лише з оборотами", value=True, key=f"act_{account}")
    sub = df.copy()
    if only_active:
        cols = [col_turn + "_a", col_turn + "_b", col_pay + "_a", col_pay + "_b"]
        sub = sub[(sub[cols].abs() > 0.5).any(axis=1)]
    sub = sub.sort_values(col_turn + "_b", ascending=False).reset_index(drop=True)

    disp = pd.DataFrame({
        "Контрагент": sub["контрагент"],
        f"{turn_label} {pa}": sub[col_turn + "_a"].map(_n),
        f"{turn_label} {pb}": sub[col_turn + "_b"].map(_n),
        "Δ": (sub[col_turn + "_b"] - sub[col_turn + "_a"]).map(_d),
        "Δ%": [_pct(a, b) for a, b in
               zip(sub[col_turn + "_a"], sub[col_turn + "_b"])],
        f"{pay_label} {pa}": sub[col_pay + "_a"].map(_n),
        f"{pay_label} {pb}": sub[col_pay + "_b"].map(_n),
        f"{debt_label} {pa}": (sub["нетто_a"] * debt_sign).map(_n),
        f"{debt_label} {pb}": (sub["нетто_b"] * debt_sign).map(_n),
    })
    st.dataframe(_style_plain(disp, ["Δ", "Δ%"]),
                 width="stretch", hide_index=True, height=_h(len(disp)))

    top = sub.head(10)
    long = pd.concat([
        pd.DataFrame({"Контрагент": top["контрагент"], "Період": pa,
                      "Сума": top[col_turn + "_a"]}),
        pd.DataFrame({"Контрагент": top["контрагент"], "Період": pb,
                      "Сума": top[col_turn + "_b"]}),
    ])
    chart = (alt.Chart(long).mark_bar()
             .encode(y=alt.Y("Контрагент:N", sort="-x", title=None),
                     x=alt.X("Сума:Q", title=f"{turn_label}, грн"),
                     yOffset="Період:N",
                     color=alt.Color("Період:N",
                                     scale=alt.Scale(range=["#94a3b8", "#1f4e78"])),
                     tooltip=["Контрагент", "Період",
                              alt.Tooltip("Сума:Q", format=",.0f")])
             .properties(height=380, title=f"Топ-10: {turn_label.lower()}"))
    st.altair_chart(chart, width="stretch")


with tab_sup:
    st.caption("Рахунок 631: **Закупівлі** — нараховано постачальниками (об. Кт), "
               "**Оплачено** — перераховано їм (об. Дт).")
    _counterparty_tab("631", col_turn="кт", col_pay="дт",
                      turn_label="Закупівлі", pay_label="Оплачено",
                      debt_label="Борг", debt_sign=-1.0)

with tab_cust:
    st.caption("Рахунок 361: **Відвантажено** — реалізація покупцям (об. Дт), "
               "**Отримано** — оплати від них (об. Кт).")
    _counterparty_tab("361", col_turn="дт", col_pay="кт",
                      turn_label="Відвантажено", pay_label="Отримано",
                      debt_label="Борг", debt_sign=1.0)

# ---------------------------------------------------------------------------
# Експорт у Excel (увесь звіт одним файлом)
# ---------------------------------------------------------------------------

def _fmt_sheet(ws, n_text_cols: int):
    """Базове оформлення: жирна шапка, ширини, формат чисел, закріплення."""
    from openpyxl.styles import Font, PatternFill
    head_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
    ws.freeze_panes = "A2"
    for i, col in enumerate(ws.columns, start=1):
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = 42 if i <= n_text_cols else 16
        if i > n_text_cols:
            for cell in col[1:]:
                cell.number_format = "#,##0"


@st.cache_data(ttl=300)
def _excel_report(pa: str, pb: str, no_internal: bool) -> bytes:
    osv = cmp.compare_osv(pa, pb)
    cash = cmp.cashflow_items(pa, pb)
    if no_internal:
        cash = cash[~cash["internal"]]
    sup = cmp.counterparty_compare(pa, pb, "631")
    cust = cmp.counterparty_compare(pa, pb, "361")

    r70, r90 = cmp.kpi_row(osv, "70"), cmp.kpi_row(osv, "90")
    r92, r93 = cmp.kpi_row(osv, "92"), cmp.kpi_row(osv, "93")
    r36 = cmp.kpi_row(osv, "36")
    fin_a, fin_b = cmp.profit_change(osv)
    money = cmp.cash_kpis(osv)
    in_a, in_b = cash["надходження_a"].sum(), cash["надходження_b"].sum()
    out_a, out_b = cash["платежі_a"].sum(), cash["платежі_b"].sum()

    kpi_rows = [
        ("Дохід (оборот Кт 70)", r70.get("об_кт_a", 0), r70.get("об_кт_b", 0)),
        ("Собівартість (оборот Дт 90)", r90.get("об_дт_a", 0), r90.get("об_дт_b", 0)),
        ("Адмін. + збут (Дт 92+93)",
         r92.get("об_дт_a", 0) + r93.get("об_дт_a", 0),
         r92.get("об_дт_b", 0) + r93.get("об_дт_b", 0)),
        ("Фінрезультат періоду (рах. 44)", fin_a, fin_b),
        ("Надходження грошей", in_a, in_b),
        ("Платежі", out_a, out_b),
        ("Гроші на кінець періоду", money["кон_a"], money["кон_b"]),
        ("Борг покупців на кінець (36)",
         r36.get("кон_нетто_a", 0), r36.get("кон_нетто_b", 0)),
    ]
    kpi = pd.DataFrame({
        "Показник": [r[0] for r in kpi_rows],
        pa: [r[1] for r in kpi_rows],
        pb: [r[2] for r in kpi_rows],
        "Δ": [r[2] - r[1] for r in kpi_rows],
        "Δ%": [_pct(r[1], r[2]) for r in kpi_rows],
    })

    osv_x = pd.DataFrame({
        "Рахунок": osv["code"],
        "Найменування": ["    " * lv + str(nm)
                         for lv, nm in zip(osv["level"], osv["name"])],
        f"Оборот Дт {pa}": osv["об_дт_a"], f"Оборот Дт {pb}": osv["об_дт_b"],
        "Δ Дт": osv["об_дт_b"] - osv["об_дт_a"],
        f"Оборот Кт {pa}": osv["об_кт_a"], f"Оборот Кт {pb}": osv["об_кт_b"],
        "Δ Кт": osv["об_кт_b"] - osv["об_кт_a"],
        f"Сальдо кін. {pa}": osv["кон_нетто_a"],
        f"Сальдо кін. {pb}": osv["кон_нетто_b"],
    })

    cash_x = pd.DataFrame({
        "Стаття": cash["стаття"],
        f"Надходження {pa}": cash["надходження_a"],
        f"Надходження {pb}": cash["надходження_b"],
        "Δ надходження": cash["надходження_b"] - cash["надходження_a"],
        f"Платежі {pa}": cash["платежі_a"],
        f"Платежі {pb}": cash["платежі_b"],
        "Δ платежі": cash["платежі_b"] - cash["платежі_a"],
    }).sort_values(f"Платежі {pb}", ascending=False)

    def _cp_x(df, turn, pay, t_lbl, p_lbl, sign):
        return pd.DataFrame({
            "Контрагент": df["контрагент"],
            f"{t_lbl} {pa}": df[turn + "_a"], f"{t_lbl} {pb}": df[turn + "_b"],
            f"Δ {t_lbl.lower()}": df[turn + "_b"] - df[turn + "_a"],
            f"{p_lbl} {pa}": df[pay + "_a"], f"{p_lbl} {pb}": df[pay + "_b"],
            f"Борг кін. {pa}": df["нетто_a"] * sign,
            f"Борг кін. {pb}": df["нетто_b"] * sign,
        }).sort_values(f"{t_lbl} {pb}", ascending=False)

    sup_x = _cp_x(sup, "кт", "дт", "Закупівлі", "Оплачено", -1.0)
    cust_x = _cp_x(cust, "дт", "кт", "Відвантажено", "Отримано", 1.0)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        for name, df, n_text in [
            ("KPI", kpi, 1), ("Рух грошей", cash_x, 1),
            ("Обороти по рахунках", osv_x, 2),
            ("Постачальники 631", sup_x, 1), ("Покупці 361", cust_x, 1),
        ]:
            df.to_excel(xw, index=False, sheet_name=name)
            _fmt_sheet(xw.sheets[name], n_text)
    return buf.getvalue()


st.divider()
st.download_button(
    f"⬇️ Завантажити звіт порівняння {pa} vs {pb} (Excel)",
    data=_excel_report(pa, pb, no_internal),
    file_name=f"порівняння_{pa}_vs_{pb}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
