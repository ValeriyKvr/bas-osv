# =============================================================================
# osv1c/excel.py — запис ОСВ (зведена + деталі) у Excel
# =============================================================================

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import config
from osv1c.connector import SummaryRow, DetailRow

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_NUM_FMT = "#,##0.00"

_SUMMARY_HEADERS = [
    "Код", "Найменування", "Рівень",
    "Сальдо поч. Дт", "Сальдо поч. Кт",
    "Оборот Дт", "Оборот Кт",
    "Сальдо кін. Дт", "Сальдо кін. Кт",
]
_DETAIL_HEADERS = [
    "Рахунок", "Код", "Субконто 1", "Субконто 2", "Субконто 3",
    "Сальдо поч. Дт", "Сальдо поч. Кт",
    "Оборот Дт", "Оборот Кт",
    "Сальдо кін. Дт", "Сальдо кін. Кт",
]


def _style_header(ws, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 50):
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, max_width)


def write_period_workbook(year_quarter_tag: str,
                          summary: list[SummaryRow],
                          details: list[DetailRow]) -> Path:
    """Записує одну книгу: аркуш «Зведена» + аркуш «Деталі»."""
    wb = openpyxl.Workbook()

    # --- Зведена ---
    ws = wb.active
    ws.title = "Зведена"
    ws.append(_SUMMARY_HEADERS)
    for r in summary:
        ws.append([
            r.account_code, r.account_name, r.level,
            r.saldo_start_dt, r.saldo_start_ct,
            r.turnover_dt, r.turnover_ct,
            r.saldo_end_dt, r.saldo_end_ct,
        ])
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=9):
        for c in row:
            c.number_format = _NUM_FMT
    _style_header(ws, len(_SUMMARY_HEADERS))
    _autosize(ws)

    # --- Деталі ---
    wd = wb.create_sheet("Деталі")
    wd.append(_DETAIL_HEADERS)
    for r in details:
        wd.append([
            r.parent_account, r.account_code,
            r.subconto1, r.subconto2, r.subconto3,
            r.saldo_start_dt, r.saldo_start_ct,
            r.turnover_dt, r.turnover_ct,
            r.saldo_end_dt, r.saldo_end_ct,
        ])
    for row in wd.iter_rows(min_row=2, min_col=6, max_col=11):
        for c in row:
            c.number_format = _NUM_FMT
    _style_header(wd, len(_DETAIL_HEADERS))
    wd.auto_filter.ref = wd.dimensions
    _autosize(wd)

    out_dir = Path(config.EXPORT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"osv_{year_quarter_tag}.xlsx"
    wb.save(out_path)
    return out_path
